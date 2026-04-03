import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

# ---- SETTINGS ----
OUTPUT_FILE = OUTPUT_FILE = f"report_{datetime.today().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
REPORT_TITLE = "Automated Data Report"

# ---- STEP 1: Load and clean data ----
def load_and_clean(path):
    df = pd.read_csv(path)

    # Auto-clean all string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip().str.title()

    df.fillna("N/A", inplace=True)
    print(f"✅ Loaded {len(df)} records from {path}")
    return df

# ---- STEP 2: Auto-detect columns ----
def detect_columns(df):
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    text_cols = df.select_dtypes(include="object").columns.tolist()

    # Pick label column: prefer common names, else first text column
    label_col = None
    for candidate in ["name", "Name", "title", "Title", "item", "Item", "id", "ID"]:
        if candidate in df.columns:
            label_col = candidate
            break
    if label_col is None and text_cols:
        label_col = text_cols[0]

    # Pick chart column: prefer common names, else first numeric column
    chart_col = None
    for candidate in ["Tasks Completed", "Sales", "Score", "Amount", "Count", "Total", "Value"]:
        if candidate in df.columns:
            chart_col = candidate
            break
    if chart_col is None and numeric_cols:
        chart_col = numeric_cols[0]

    print(f"   Label column  : {label_col}")
    print(f"   Numeric columns: {numeric_cols}")
    print(f"   Chart column  : {chart_col}")

    return label_col, numeric_cols, chart_col

# ---- STEP 3: Auto-generate summary stats ----
def generate_summary(df, numeric_cols):
    summary = {"Total Records": len(df)}

    for col in numeric_cols:
        summary[f"Avg {col}"] = round(df[col].mean(), 1)
        summary[f"Max {col}"] = df[col].max()
        summary[f"Min {col}"] = df[col].min()

    # Count unique values in text columns (useful for categories)
    for col in df.select_dtypes(include="object").columns:
        unique_vals = df[col].nunique()
        if 2 <= unique_vals <= 10:  # only useful categorical columns
            summary[f"Unique {col}"] = unique_vals

    return summary

# ---- STEP 4: Write Excel report ----
def write_excel(df, summary, label_col, chart_col, output_path, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Styles ---
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill    = PatternFill("solid", fgColor="1F4E79")
    title_font     = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    summary_font   = Font(name="Calibri", bold=True, size=11)
    center         = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # --- Title ---
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated on: {datetime.today().strftime('%d %B %Y')}"
    ws["A2"].alignment = center
    ws["A2"].font = Font(italic=True, color="888888")

    # --- Summary Section ---
    ws["A4"] = "SUMMARY"
    ws["A4"].font = Font(bold=True, size=12, color="1F4E79")

    row = 5
    for key, value in summary.items():
        ws.cell(row=row, column=1, value=key).font = summary_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    # --- Data Table ---
    table_start_row = row + 2
    ws.cell(row=table_start_row, column=1, value="DETAILED DATA").font = Font(bold=True, size=12, color="1F4E79")
    table_start_row += 1

    columns = list(df.columns)
    for col_num, col_name in enumerate(columns, 1):
        cell = ws.cell(row=table_start_row, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for row_num, row_data in enumerate(df.itertuples(index=False), table_start_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # --- Auto column width ---
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # --- Bar Chart (auto-detected column) ---
    if label_col and chart_col:
        chart_sheet = wb.create_sheet("Chart")
        chart_sheet["A1"] = label_col
        chart_sheet["B1"] = chart_col

        for i, (label, value) in enumerate(zip(df[label_col], df[chart_col]), 2):
            chart_sheet[f"A{i}"] = label
            chart_sheet[f"B{i}"] = value

        chart = BarChart()
        chart.type = "col"
        chart.title = f"{chart_col} by {label_col}"
        chart.y_axis.title = chart_col
        chart.x_axis.title = label_col
        chart.style = 10

        data_ref = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(df) + 1)
        cats_ref = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(df) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 20
        chart.height = 12
        chart_sheet.add_chart(chart, "D2")
    else:
        print("⚠️  No suitable columns found for chart — skipping chart sheet.")

    wb.save(output_path)
    print(f"💾 Report saved as: {output_path}")

# ---- MAIN ----
def main():
    print("🚀 Starting report generation...\n")

    INPUT_FILE = input("📂 Enter CSV filename (e.g. data.csv): ").strip()

    # --- File + column check ---
    try:
        preview = pd.read_csv(INPUT_FILE, nrows=0)
        print(f"✅ File found. Columns detected: {list(preview.columns)}\n")
    except FileNotFoundError:
        print(f"\n❌ File not found: '{INPUT_FILE}'")
        print("   Make sure the file is in the same folder as this script.")
        return

    df = load_and_clean(INPUT_FILE)

    print("\n🔍 Auto-detecting columns...")
    label_col, numeric_cols, chart_col = detect_columns(df)

    if not numeric_cols:
        print("\n⚠️  No numeric columns found. Summary and chart will be limited.")

    summary = generate_summary(df, numeric_cols)

    print("\n📊 Summary:")
    for k, v in summary.items():
        print(f"   {k}: {v}")

    print("\n📝 Writing Excel report...")
    write_excel(df, summary, label_col, chart_col, OUTPUT_FILE, REPORT_TITLE)
    print("\n✅ Done! Open your report to see the result.")

main()